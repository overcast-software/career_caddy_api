from io import BytesIO
from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework import status
from openpyxl import load_workbook
from job_hunting.models import Company, JobPost, JobApplication, Question, Answer

User = get_user_model()


class TestCareerDataExport(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_superuser(
            username="exportuser", password="pass", email="e@x.com"
        )
        self.client.force_authenticate(user=self.user)
        self.company = Company.objects.create(name="Acme")
        self.jp = JobPost.objects.create(
            title="Engineer", company=self.company, link="https://acme.com/1",
            created_by=self.user,
        )
        self.ja = JobApplication.objects.create(
            user=self.user, job_post=self.jp, company=self.company, status="Applied",
        )
        self.q = Question.objects.create(
            application=self.ja, company=self.company, created_by=self.user,
            content="Why Acme?", favorite=True,
        )
        self.a = Answer.objects.create(
            question=self.q, content="Great culture", favorite=False, status="draft",
        )

    def test_export_returns_xlsx(self):
        resp = self.client.get("/api/v1/career-data/export/")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        self.assertIn("attachment", resp["Content-Disposition"])

    def test_export_contains_all_sheets(self):
        resp = self.client.get("/api/v1/career-data/export/")
        wb = load_workbook(BytesIO(resp.content))
        self.assertEqual(set(wb.sheetnames), {"job-posts", "job-applications", "questions", "answers"})

    def test_export_data_integrity(self):
        resp = self.client.get("/api/v1/career-data/export/")
        wb = load_workbook(BytesIO(resp.content))
        # job-posts: header + 1 row
        jp_rows = list(wb["job-posts"].iter_rows(values_only=True))
        self.assertEqual(len(jp_rows), 2)
        self.assertEqual(jp_rows[1][1], "Engineer")  # title column
        # questions
        q_rows = list(wb["questions"].iter_rows(values_only=True))
        self.assertEqual(len(q_rows), 2)
        self.assertEqual(q_rows[1][4], "Why Acme?")  # content column

    def test_export_requires_auth(self):
        anon = APIClient()
        resp = anon.get("/api/v1/career-data/export/")
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)


class TestCareerDataImport(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.superuser = User.objects.create_superuser(
            username="importsuper", password="pass", email="s@x.com"
        )
        self.regular = User.objects.create_user(username="importreg", password="pass")
        self.client.force_authenticate(user=self.superuser)

    def _make_xlsx(self):
        """Build a minimal xlsx in memory with one record per sheet."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "job-posts"
        ws.append(["id", "title", "company", "description", "link",
                    "posted_date", "extraction_date", "salary_min", "salary_max",
                    "location", "remote", "created_at"])
        ws.append([1, "Dev", "NewCo", "Build stuff", "https://newco.com/dev",
                    None, None, None, None, "Remote", True, None])

        ws2 = wb.create_sheet("job-applications")
        ws2.append(["id", "job_post_id", "company", "status",
                     "applied_at", "tracking_url", "notes"])
        ws2.append([1, 1, "NewCo", "Applied", None, None, "Great fit"])

        ws3 = wb.create_sheet("questions")
        ws3.append(["id", "application_id", "company", "job_post_id", "content", "favorite", "created_at"])
        ws3.append([1, 1, "NewCo", 1, "Tell me about yourself", True, None])

        ws4 = wb.create_sheet("answers")
        ws4.append(["id", "question_id", "content", "favorite", "status", "created_at"])
        ws4.append([1, 1, "I'm a developer", False, "final", None])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "test.xlsx"
        return buf

    def test_import_creates_records(self):
        f = self._make_xlsx()
        resp = self.client.post("/api/v1/career-data/import/", {"file": f}, format="multipart")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()["data"]
        self.assertEqual(data["job-posts"]["created"], 1)
        self.assertEqual(data["job-applications"]["created"], 1)
        self.assertEqual(data["questions"]["created"], 1)
        self.assertEqual(data["answers"]["created"], 1)
        self.assertEqual(JobPost.objects.count(), 1)
        self.assertEqual(Question.objects.count(), 1)

    def test_import_records_discovery_for_caller(self):
        """CSV import must record JobPostDiscovery for the importer.
        Without this, imported posts have zero per-user signals and
        are invisible until the user manually triggers another action."""
        from job_hunting.models import JobPostDiscovery

        f = self._make_xlsx()
        self.client.post("/api/v1/career-data/import/", {"file": f}, format="multipart")
        jp = JobPost.objects.get(link="https://newco.com/dev")
        disc = JobPostDiscovery.objects.filter(
            job_post=jp, user=self.superuser
        ).first()
        self.assertIsNotNone(disc, "import must record discovery for caller")
        self.assertEqual(disc.source, "import")

    def test_import_dedupe_merges_empty_company(self):
        """A pre-existing JobPost with the same link but NULL company
        gets its company filled from the imported row. Same merge policy
        as the create() endpoint."""
        from job_hunting.models import JobPostDiscovery

        existing = JobPost.objects.create(
            title="Stub",
            company=None,
            link="https://newco.com/dev",
            description="x" * 500,
            created_by=self.regular,
        )
        f = self._make_xlsx()
        self.client.post("/api/v1/career-data/import/", {"file": f}, format="multipart")
        existing.refresh_from_db()
        self.assertIsNotNone(
            existing.company_id,
            "import must merge empty company_id onto existing duplicate",
        )
        self.assertEqual(existing.company.name, "NewCo")
        # And the importer gets a discovery row even on the dedupe path.
        self.assertTrue(
            JobPostDiscovery.objects.filter(
                job_post=existing, user=self.superuser, source="import"
            ).exists()
        )

    def test_import_skips_duplicates(self):
        f1 = self._make_xlsx()
        self.client.post("/api/v1/career-data/import/", {"file": f1}, format="multipart")
        f2 = self._make_xlsx()
        resp = self.client.post("/api/v1/career-data/import/", {"file": f2}, format="multipart")
        data = resp.json()["data"]
        self.assertEqual(data["job-posts"]["skipped"], 1)
        self.assertEqual(data["job-posts"]["created"], 0)
        self.assertEqual(data["job-applications"]["skipped"], 1)
        self.assertEqual(data["questions"]["skipped"], 1)
        self.assertEqual(data["answers"]["skipped"], 1)
        self.assertEqual(JobPost.objects.count(), 1)

    def test_import_skips_duplicates_without_link(self):
        """Job posts without a link should still be deduplicated by title+company+user."""
        from openpyxl import Workbook

        def _make_no_link_xlsx():
            wb = Workbook()
            ws = wb.active
            ws.title = "job-posts"
            ws.append(["id", "title", "company", "description", "link",
                        "posted_date", "extraction_date", "salary_min", "salary_max",
                        "location", "remote", "created_at"])
            ws.append([1, "Manual Entry", "SomeCo", "No link job", None,
                        None, None, None, None, "Remote", True, None])

            ws2 = wb.create_sheet("job-applications")
            ws2.append(["id", "job_post_id", "company", "status",
                         "applied_at", "tracking_url", "notes"])
            ws2.append([1, 1, "SomeCo", "Applied", None, None, "Notes"])

            ws3 = wb.create_sheet("questions")
            ws3.append(["id", "application_id", "company", "job_post_id",
                         "content", "favorite", "created_at"])
            ws3.append([1, 1, "SomeCo", 1, "Why here?", False, None])

            ws4 = wb.create_sheet("answers")
            ws4.append(["id", "question_id", "content", "favorite", "status", "created_at"])
            ws4.append([1, 1, "Because reasons", False, "draft", None])

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            buf.name = "nolink.xlsx"
            return buf

        # First import
        f1 = _make_no_link_xlsx()
        resp1 = self.client.post("/api/v1/career-data/import/", {"file": f1}, format="multipart")
        self.assertEqual(resp1.json()["data"]["job-posts"]["created"], 1)
        self.assertEqual(JobPost.objects.count(), 1)

        # Second import — should skip everything
        f2 = _make_no_link_xlsx()
        resp2 = self.client.post("/api/v1/career-data/import/", {"file": f2}, format="multipart")
        data = resp2.json()["data"]
        self.assertEqual(data["job-posts"]["skipped"], 1)
        self.assertEqual(data["job-posts"]["created"], 0)
        self.assertEqual(data["job-applications"]["skipped"], 1)
        self.assertEqual(data["questions"]["skipped"], 1)
        self.assertEqual(data["answers"]["skipped"], 1)
        self.assertEqual(JobPost.objects.count(), 1)

    def test_import_superuser_only(self):
        self.client.force_authenticate(user=self.regular)
        f = self._make_xlsx()
        resp = self.client.post("/api/v1/career-data/import/", {"file": f}, format="multipart")
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

    def test_import_no_file_400(self):
        resp = self.client.post("/api/v1/career-data/import/", {}, format="multipart")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_import_bad_file_400(self):
        bad = BytesIO(b"not an xlsx")
        bad.name = "bad.xlsx"
        resp = self.client.post("/api/v1/career-data/import/", {"file": bad}, format="multipart")
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)


class TestExportImportRoundTrip(TestCase):
    """Export data, clear DB, import it back — verify integrity."""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_superuser(
            username="roundtrip", password="pass", email="r@x.com"
        )
        self.client.force_authenticate(user=self.user)

    def test_round_trip(self):
        co = Company.objects.create(name="RoundCo")
        jp = JobPost.objects.create(
            title="RoundDev", company=co, link="https://round.co/1", created_by=self.user,
        )
        ja = JobApplication.objects.create(
            user=self.user, job_post=jp, company=co, status="Interviewing",
        )
        q = Question.objects.create(
            application=ja, company=co, created_by=self.user,
            content="Round question?", favorite=True,
        )
        Answer.objects.create(question=q, content="Round answer", favorite=True, status="final")

        # Export
        resp = self.client.get("/api/v1/career-data/export/")
        xlsx_bytes = resp.content

        # Clear
        Answer.objects.all().delete()
        Question.objects.all().delete()
        JobApplication.objects.all().delete()
        JobPost.objects.all().delete()
        self.assertEqual(JobPost.objects.count(), 0)

        # Import
        f = BytesIO(xlsx_bytes)
        f.name = "roundtrip.xlsx"
        resp = self.client.post("/api/v1/career-data/import/", {"file": f}, format="multipart")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()["data"]
        self.assertEqual(data["job-posts"]["created"], 1)
        self.assertEqual(data["job-applications"]["created"], 1)
        self.assertEqual(data["questions"]["created"], 1)
        self.assertEqual(data["answers"]["created"], 1)

        # Verify
        self.assertEqual(JobPost.objects.first().title, "RoundDev")
        self.assertEqual(Question.objects.first().content, "Round question?")
        self.assertTrue(Answer.objects.first().favorite)


class TestExportIsolatesUsers(TestCase):
    """BACK-133 regression: the export must never cross user boundaries.

    The endpoint was gated on IsAuthenticated with four completely unfiltered
    querysets, so any authenticated caller downloaded every user's job
    applications, notes, questions and answers.

    It survived because every pre-existing test in this file runs a single
    superuser fixture: with one user in the database, an unscoped queryset and
    a correctly scoped one return identical rows. Only a second user makes the
    two distinguishable, which is what this class adds.
    """

    def setUp(self):
        self.client = APIClient()
        self.mine = User.objects.create_user(username="mine", password="pass")
        self.theirs = User.objects.create_user(username="theirs", password="pass")

        self.company = Company.objects.create(name="Shared Co")

        # Mine
        self.my_jp = JobPost.objects.create(
            title="MY ROLE", company=self.company,
            link="https://shared.co/mine", created_by=self.mine,
        )
        self.my_ja = JobApplication.objects.create(
            user=self.mine, job_post=self.my_jp, company=self.company,
            status="Applied", notes="MY PRIVATE NOTE",
        )
        self.my_q = Question.objects.create(
            application=self.my_ja, company=self.company, created_by=self.mine,
            content="MY QUESTION", favorite=True,
        )
        Answer.objects.create(
            question=self.my_q, content="MY ANSWER", favorite=True, status="draft",
        )

        # Theirs — same company, so a company join cannot be what separates them
        self.their_jp = JobPost.objects.create(
            title="THEIR ROLE", company=self.company,
            link="https://shared.co/theirs", created_by=self.theirs,
        )
        self.their_ja = JobApplication.objects.create(
            user=self.theirs, job_post=self.their_jp, company=self.company,
            status="Applied", notes="THEIR PRIVATE NOTE",
        )
        self.their_q = Question.objects.create(
            application=self.their_ja, company=self.company, created_by=self.theirs,
            content="THEIR QUESTION", favorite=True,
        )
        Answer.objects.create(
            question=self.their_q, content="THEIR ANSWER", favorite=True, status="draft",
        )

    def _export_text(self, user):
        self.client.force_authenticate(user=user)
        resp = self.client.get("/api/v1/career-data/export/")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        wb = load_workbook(BytesIO(resp.content))
        chunks = []
        for name in wb.sheetnames:
            for row in wb[name].iter_rows(values_only=True):
                chunks.append(" ".join("" if c is None else str(c) for c in row))
        return "\n".join(chunks)

    def test_export_excludes_other_users_rows(self):
        text = self._export_text(self.mine)
        self.assertIn("MY ROLE", text)
        self.assertIn("MY PRIVATE NOTE", text)
        self.assertIn("MY QUESTION", text)
        self.assertIn("MY ANSWER", text)
        for leaked in ("THEIR ROLE", "THEIR PRIVATE NOTE", "THEIR QUESTION", "THEIR ANSWER"):
            self.assertNotIn(leaked, text, f"{leaked} leaked into another user's export")

    def test_export_is_symmetric(self):
        text = self._export_text(self.theirs)
        self.assertIn("THEIR ROLE", text)
        for leaked in ("MY ROLE", "MY PRIVATE NOTE", "MY QUESTION", "MY ANSWER"):
            self.assertNotIn(leaked, text, f"{leaked} leaked into another user's export")

    def test_staff_get_no_cross_user_export(self):
        """Doug, 2026-08-27: 'No cross user export.'

        JobPostQuerySet.visible_to returns the whole table for staff. The
        export deliberately does not use it — a staff account exports its own
        data and nothing else. If a cross-user export is ever wanted it needs
        an explicit user_id, not a silent everything.
        """
        staff = User.objects.create_superuser(
            username="staff", password="pass", email="s@x.com"
        )
        text = self._export_text(staff)
        for leaked in ("MY ROLE", "THEIR ROLE", "MY PRIVATE NOTE", "THEIR PRIVATE NOTE"):
            self.assertNotIn(leaked, text, f"staff export leaked {leaked}")


class TestCareerDataExportIllegalCharacters(TestCase):
    """BACK-137: one control character anywhere 500s the whole export.

    openpyxl raises IllegalCharacterError for the control chars its own
    ``ILLEGAL_CHARACTERS_RE`` matches (0x00-0x08, 0x0B-0x0C, 0x0E-0x1F), so a
    single bad byte in one row killed ``GET /api/v1/career-data/export/``
    outright. It was found in prod in an email-sourced JobPost.description,
    where a \\x0b arrived via the cc_auto triage path's quoted-printable /
    HTML-to-text conversion.

    The old fixtures could not catch this: none of them contained a control
    character. These do -- and setUp deliberately keeps them OUT, so each
    per-sheet test dirties only the column it asserts on. Otherwise a control
    char in the first sheet written 500s the export and every test in the
    class fails together, which would make the per-sheet names a lie.
    """

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username="ctrlcharuser", password="pass")
        self.client.force_authenticate(user=self.user)
        self.company = Company.objects.create(name="Acme")
        self.jp = JobPost.objects.create(
            title="Engineer Role",
            company=self.company,
            link="https://acme.com/1",
            description="Source: direct email from A Recruiter See JD below",
            created_by=self.user,
        )
        self.ja = JobApplication.objects.create(
            user=self.user,
            job_post=self.jp,
            company=self.company,
            status="Applied",
            notes="Phone screen went well",
        )
        self.q = Question.objects.create(
            application=self.ja,
            company=self.company,
            created_by=self.user,
            content="Why Acme?",
            favorite=True,
        )
        self.a = Answer.objects.create(
            question=self.q,
            content="Great culture",
            favorite=False,
            status="draft",
        )

    def _workbook(self):
        resp = self.client.get("/api/v1/career-data/export/")
        self.assertEqual(
            resp.status_code,
            status.HTTP_200_OK,
            "export must not 500 on control characters in the data",
        )
        return load_workbook(BytesIO(resp.content))

    def _row(self, sheet):
        return list(self._workbook()[sheet].iter_rows(values_only=True))[1]

    def test_export_survives_control_characters(self):
        # The one test that dirties every sheet at once: the bare fact that
        # this returns a readable workbook is the regression -- before the fix
        # the view raised IllegalCharacterError on the first bad cell.
        self.jp.title = "Engineer\x0cRole"
        self.jp.description = "Source: direct email from A Recruiter\x0bSee JD below"
        self.jp.save(update_fields=["title", "description"])
        self.ja.notes = "Phone screen\x01 went well"
        self.ja.save(update_fields=["notes"])
        self.q.content = "Why Acme?\x1f"
        self.q.save(update_fields=["content"])
        self.a.content = "Great\x08 culture"
        self.a.save(update_fields=["content"])

        wb = self._workbook()
        self.assertEqual(
            set(wb.sheetnames),
            {"job-posts", "job-applications", "questions", "answers"},
        )

    def test_job_posts_sheet_is_scrubbed(self):
        # Shaped like the row that actually blew up in prod.
        self.jp.title = "Engineer\x0cRole"
        self.jp.description = "Source: direct email from A Recruiter\x0bSee JD below"
        self.jp.save(update_fields=["title", "description"])
        row = self._row("job-posts")
        self.assertEqual(row[1], "Engineer\nRole")
        self.assertEqual(row[3], "Source: direct email from A Recruiter\nSee JD below")

    def test_vertical_tab_and_form_feed_become_newlines(self):
        """VT/FF are line and page breaks in the source, not noise.

        Deleting them jams the words on either side together
        ("...A RecruiterSee JD below"), silently destroying a word boundary in
        the exported cell. They map to "\\n" instead, which openpyxl accepts.
        """
        self.jp.description = "before\x0bafter\x0cend"
        self.jp.save(update_fields=["description"])
        row = self._row("job-posts")
        self.assertEqual(row[3], "before\nafter\nend")
        self.assertNotIn("beforeafter", row[3])

    def test_job_applications_sheet_is_scrubbed(self):
        self.ja.notes = "Phone screen\x01 went well"
        self.ja.save(update_fields=["notes"])
        self.assertEqual(self._row("job-applications")[6], "Phone screen went well")

    def test_questions_sheet_is_scrubbed(self):
        self.q.content = "Why Acme?\x1f"
        self.q.save(update_fields=["content"])
        self.assertEqual(self._row("questions")[4], "Why Acme?")

    def test_answers_sheet_is_scrubbed(self):
        self.a.content = "Great\x08 culture"
        self.a.save(update_fields=["content"])
        self.assertEqual(self._row("answers")[2], "Great culture")

    def test_legal_whitespace_is_preserved(self):
        """Tab / newline / CR are legal and carry real meaning in a pasted job
        description -- over-scrubbing would silently mangle every multi-line
        cell in the export. CR is asserted too: openpyxl writes it as the
        numeric reference ``&#13;``, which is exempt from the XML line-ending
        normalisation that would otherwise collapse it into the "\\n"."""
        self.jp.description = "Line one\nLine two\tindented\r\nLine three"
        self.jp.save(update_fields=["description"])
        row = self._row("job-posts")
        self.assertEqual(row[3], "Line one\nLine two\tindented\r\nLine three")
        self.assertIn("\n", row[3])
        self.assertIn("\t", row[3])
        self.assertIn("\r", row[3])
