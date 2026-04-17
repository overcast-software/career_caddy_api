from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from drf_spectacular.utils import (
    extend_schema,
    OpenApiResponse,
    inline_serializer,
    OpenApiParameter,
)
from drf_spectacular.types import OpenApiTypes
from rest_framework import serializers as drf_serializers

from job_hunting.lib.services.application_prompt_builder import ApplicationPromptBuilder
from job_hunting.lib.services.answer_service import AnswerService
from job_hunting.lib.models import CareerData
from job_hunting.models import (
    Question,
    JobPost,
    Answer,
    JobApplication,
    Resume,
)


@extend_schema(
    tags=["Career Data"],
    summary="Get aggregated career data as an AI-ready prompt string",
    parameters=[
        OpenApiParameter(
            "user_id",
            OpenApiTypes.INT,
            OpenApiParameter.PATH,
            required=False,
            description="Target user ID (defaults to authenticated user)",
        ),
    ],
    responses={
        200: OpenApiResponse(
            description="Career data formatted as a prompt string",
            response=inline_serializer(
                name="CareerDataResponse",
                fields={"data": drf_serializers.CharField()},
            ),
        )
    },
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def career_data(request, user_id=None):
    # Get aggregated career data for the authenticated user or specified user (with API key).
    # Determine which user's data to return
    target_user_id = user_id if user_id is not None else request.user.id

    # If accessing another user's data, ensure proper authorization
    if user_id is not None and user_id != request.user.id:
        if not request.user.is_staff:
            return Response({"errors": [{"detail": "Forbidden"}]}, status=403)

    career_data = CareerData.for_user(target_user_id)

    prompt_builder = ApplicationPromptBuilder(max_section_chars=60000)
    career_data_prompt = prompt_builder.build_from_career_data(career_data)
    return Response({"data": career_data_prompt, "meta": career_data.to_refs()})


@extend_schema(
    tags=["Career Data"],
    summary="Generate an AI prompt for answering a job application question",
    request=inline_serializer(
        name="GeneratePromptRequest",
        fields={
            "question_id": drf_serializers.IntegerField(
                help_text="Required. ID of the Question to answer."
            ),
            "job_post_id": drf_serializers.IntegerField(
                required=False, help_text="Optional job post context."
            ),
            "resume_id": drf_serializers.IntegerField(
                required=False, help_text="Optional resume to include."
            ),
            "instructions": drf_serializers.CharField(
                required=False, help_text="Custom instructions appended to the prompt."
            ),
        },
    ),
    responses={
        200: OpenApiResponse(
            description="Generated prompt and context metadata",
            response=inline_serializer(
                name="GeneratePromptResponse",
                fields={
                    "data": inline_serializer(
                        name="GeneratePromptData",
                        fields={
                            "prompt": drf_serializers.CharField(),
                            "context": drf_serializers.DictField(),
                        },
                    )
                },
            ),
        ),
        400: OpenApiResponse(
            description="Missing or invalid question_id / job_post_id / resume_id"
        ),
        403: OpenApiResponse(description="Resume not accessible"),
        404: OpenApiResponse(description="Question not found"),
    },
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def generate_prompt(request):
    """
    Generate an AI prompt using ApplicationPromptBuilder for a specific question.

    Expects JSON payload with:
    - question_id: ID of the question to answer (required)
    - job_post_id: (optional) ID of the job post
    - resume_id: (optional) ID of the resume to use
    - instructions: (optional) Custom instructions for the prompt
    """
    data = request.data if isinstance(request.data, dict) else {}

    # Extract required question_id
    question_id = data.get("question_id")
    if not question_id:
        return Response(
            {"errors": [{"detail": "question_id is required"}]},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        question_id = int(question_id)
        question = Question.objects.filter(pk=question_id).first()
        if not question:
            return Response(
                {"errors": [{"detail": "Question not found"}]},
                status=status.HTTP_404_NOT_FOUND,
            )
    except (TypeError, ValueError):
        return Response(
            {"errors": [{"detail": "Invalid question_id"}]},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Extract optional parameters
    job_post_id = data.get("job_post_id")
    resume_id = data.get("resume_id")
    instructions = data.get("instructions")

    # Use AnswerService to load comprehensive context data

    answer_service = AnswerService(
        ai_client=None
    )  # No AI client needed for data aggregation
    context = answer_service.load_context_for_question(question)

    # Override context with specific parameters if provided
    if job_post_id:
        try:
            job_post_id = int(job_post_id)
            job_post = JobPost.objects.filter(pk=job_post_id).first()
            if job_post:
                context["job_post"] = job_post
                if hasattr(job_post, "company"):
                    context["company"] = job_post.company
        except (TypeError, ValueError):
            return Response(
                {"errors": [{"detail": "Invalid job_post_id"}]},
                status=status.HTTP_400_BAD_REQUEST,
            )

    if resume_id:
        try:
            resume_id = int(resume_id)
            resume = Resume.get(resume_id)
            if resume and resume.user_id == request.user.id:
                context["resume"] = resume
                context["resumes"] = [resume]
            elif resume and resume.user_id != request.user.id:
                return Response(
                    {"errors": [{"detail": "Resume not accessible"}]},
                    status=status.HTTP_403_FORBIDDEN,
                )
        except (TypeError, ValueError):
            return Response(
                {"errors": [{"detail": "Invalid resume_id"}]},
                status=status.HTTP_400_BAD_REQUEST,
            )

    # Create prompt builder and generate prompt
    builder = ApplicationPromptBuilder()
    prompt = builder.build(context, instructions)

    return Response(
        {
            "data": {
                "prompt": prompt,
                "context": {
                    "question_id": question.id,
                    "job_post_id": (
                        context.get("job_post").id if context.get("job_post") else None
                    ),
                    "resume_ids": [r.id for r in context.get("resumes", [])],
                    "cover_letter_count": len(context.get("cover_letters", [])),
                    "qa_count": len(context.get("qas", [])),
                },
            }
        }
    )


# ---------------------------------------------------------------------------
# Career-data export / import
# ---------------------------------------------------------------------------


@extend_schema(
    tags=["Career Data"],
    summary="Export career data as an Excel (.xlsx) file",
    responses={
        200: OpenApiResponse(
            description="Excel workbook with sheets: job-posts, job-applications, questions, answers"
        )
    },
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def career_data_export(request):
    from openpyxl import Workbook

    wb = Workbook()

    # -- job-posts --
    ws = wb.active
    ws.title = "job-posts"
    jp_headers = [
        "id", "title", "company", "description", "link",
        "posted_date", "extraction_date", "salary_min", "salary_max",
        "location", "remote", "created_at",
    ]
    ws.append(jp_headers)
    for jp in JobPost.objects.select_related("company").order_by("id"):
        ws.append([
            jp.id, jp.title,
            jp.company.name if jp.company else None,
            jp.description, jp.link,
            str(jp.posted_date) if jp.posted_date else None,
            str(jp.extraction_date) if jp.extraction_date else None,
            float(jp.salary_min) if jp.salary_min is not None else None,
            float(jp.salary_max) if jp.salary_max is not None else None,
            jp.location,
            jp.remote,
            jp.created_at.isoformat() if jp.created_at else None,
        ])

    # -- job-applications --
    ws2 = wb.create_sheet("job-applications")
    ja_headers = [
        "id", "job_post_id", "company", "status",
        "applied_at", "tracking_url", "notes",
    ]
    ws2.append(ja_headers)
    for ja in JobApplication.objects.select_related("company").order_by("id"):
        ws2.append([
            ja.id, ja.job_post_id,
            ja.company.name if ja.company else None,
            ja.status,
            ja.applied_at.isoformat() if ja.applied_at else None,
            ja.tracking_url, ja.notes,
        ])

    # -- questions --
    ws3 = wb.create_sheet("questions")
    q_headers = ["id", "application_id", "company", "job_post_id", "content", "favorite", "created_at"]
    ws3.append(q_headers)
    for q in Question.objects.select_related("company").order_by("id"):
        ws3.append([
            q.id, q.application_id,
            q.company.name if q.company else None,
            q.job_post_id, q.content, q.favorite,
            q.created_at.isoformat() if q.created_at else None,
        ])

    # -- answers --
    ws4 = wb.create_sheet("answers")
    a_headers = ["id", "question_id", "content", "favorite", "status", "created_at"]
    ws4.append(a_headers)
    for a in Answer.objects.order_by("id"):
        ws4.append([
            a.id, a.question_id, a.content, a.favorite, a.status,
            a.created_at.isoformat() if a.created_at else None,
        ])

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="career-caddy-export.xlsx"'
    return response


@extend_schema(
    tags=["Career Data"],
    summary="Import career data from an Excel (.xlsx) file",
    request={"multipart/form-data": {"type": "object", "properties": {"file": {"type": "string", "format": "binary"}}}},
    responses={
        200: OpenApiResponse(description="Import summary with created/skipped counts"),
        400: OpenApiResponse(description="Missing or invalid file"),
        403: OpenApiResponse(description="Superuser access required"),
    },
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def career_data_import(request):
    if not request.user.is_superuser:
        return Response(
            {"errors": [{"detail": "Superuser access required"}]},
            status=status.HTTP_403_FORBIDDEN,
        )

    uploaded = request.FILES.get("file")
    if not uploaded:
        return Response(
            {"errors": [{"detail": "No file provided. Upload an .xlsx file as 'file'."}]},
            status=status.HTTP_400_BAD_REQUEST,
        )

    from openpyxl import load_workbook
    from django.db import transaction
    from datetime import datetime

    try:
        wbook = load_workbook(uploaded, read_only=True)
    except Exception:
        return Response(
            {"errors": [{"detail": "Could not read file as .xlsx"}]},
            status=status.HTTP_400_BAD_REQUEST,
        )

    stats = {"job-posts": {"created": 0, "skipped": 0},
             "job-applications": {"created": 0, "skipped": 0},
             "questions": {"created": 0, "skipped": 0},
             "answers": {"created": 0, "skipped": 0}}

    def _parse_datetime(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        try:
            return datetime.fromisoformat(str(val))
        except (ValueError, TypeError):
            return None

    def _parse_date(val):
        if val is None:
            return None
        from datetime import date
        if isinstance(val, date):
            return val
        try:
            return date.fromisoformat(str(val))
        except (ValueError, TypeError):
            return None

    def _rows_as_dicts(sheet):
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return
        headers = [str(h).strip() if h else "" for h in headers]
        for row in rows:
            yield dict(zip(headers, row))

    with transaction.atomic():
        # Map old IDs → new IDs for relational integrity
        jp_id_map = {}  # old job_post id → new job_post id
        ja_id_map = {}  # old application id → new application id
        q_id_map = {}   # old question id → new question id

        # -- job-posts --
        if "job-posts" in wbook.sheetnames:
            from job_hunting.models import Company
            for row in _rows_as_dicts(wbook["job-posts"]):
                old_id = row.get("id")
                link = row.get("link")
                company = None
                if row.get("company"):
                    company, _ = Company.objects.get_or_create(name=row["company"])
                # Skip duplicate by link (unique constraint)
                existing = None
                if link:
                    existing = JobPost.objects.filter(link=link).first()
                # Fallback: match on title + company + created_by for linkless posts
                if not existing and row.get("title"):
                    existing = JobPost.objects.filter(
                        title=row["title"], company=company, created_by=request.user
                    ).first()
                if existing:
                    if old_id is not None:
                        jp_id_map[int(old_id)] = existing.id
                    stats["job-posts"]["skipped"] += 1
                    continue
                from decimal import Decimal
                jp = JobPost.objects.create(
                    title=row.get("title"),
                    company=company,
                    description=row.get("description"),
                    link=link,
                    posted_date=_parse_date(row.get("posted_date")),
                    extraction_date=_parse_date(row.get("extraction_date")),
                    salary_min=Decimal(str(row["salary_min"])) if row.get("salary_min") is not None else None,
                    salary_max=Decimal(str(row["salary_max"])) if row.get("salary_max") is not None else None,
                    location=row.get("location"),
                    remote=row.get("remote"),
                    created_by=request.user,
                )
                if old_id is not None:
                    jp_id_map[int(old_id)] = jp.id
                stats["job-posts"]["created"] += 1

        # -- job-applications --
        if "job-applications" in wbook.sheetnames:
            from job_hunting.models import Company
            for row in _rows_as_dicts(wbook["job-applications"]):
                old_id = row.get("id")
                old_jp_id = row.get("job_post_id")
                new_jp_id = jp_id_map.get(int(old_jp_id)) if old_jp_id is not None else None
                # Skip duplicate: same user + same job_post
                if new_jp_id and JobApplication.objects.filter(
                    user=request.user, job_post_id=new_jp_id
                ).exists():
                    if old_id is not None:
                        existing = JobApplication.objects.filter(
                            user=request.user, job_post_id=new_jp_id
                        ).first()
                        ja_id_map[int(old_id)] = existing.id
                    stats["job-applications"]["skipped"] += 1
                    continue
                company = None
                if row.get("company"):
                    company, _ = Company.objects.get_or_create(name=row["company"])
                ja = JobApplication.objects.create(
                    user=request.user,
                    job_post_id=new_jp_id,
                    company=company,
                    status=row.get("status"),
                    applied_at=_parse_datetime(row.get("applied_at")),
                    tracking_url=row.get("tracking_url"),
                    notes=row.get("notes"),
                )
                if old_id is not None:
                    ja_id_map[int(old_id)] = ja.id
                stats["job-applications"]["created"] += 1

        # -- questions --
        if "questions" in wbook.sheetnames:
            from job_hunting.models import Company
            for row in _rows_as_dicts(wbook["questions"]):
                old_id = row.get("id")
                old_app_id = row.get("application_id")
                new_app_id = ja_id_map.get(int(old_app_id)) if old_app_id is not None else None
                old_jp_id = row.get("job_post_id")
                new_jp_id = jp_id_map.get(int(old_jp_id)) if old_jp_id is not None else None
                content = row.get("content")
                # Skip duplicate: same content + same application
                if content and new_app_id and Question.objects.filter(
                    content=content, application_id=new_app_id
                ).exists():
                    if old_id is not None:
                        existing = Question.objects.filter(
                            content=content, application_id=new_app_id
                        ).first()
                        q_id_map[int(old_id)] = existing.id
                    stats["questions"]["skipped"] += 1
                    continue
                company = None
                if row.get("company"):
                    company, _ = Company.objects.get_or_create(name=row["company"])
                q = Question.objects.create(
                    application_id=new_app_id,
                    company=company,
                    created_by=request.user,
                    job_post_id=new_jp_id,
                    content=content,
                    favorite=bool(row.get("favorite")),
                )
                if old_id is not None:
                    q_id_map[int(old_id)] = q.id
                stats["questions"]["created"] += 1

        # -- answers --
        if "answers" in wbook.sheetnames:
            for row in _rows_as_dicts(wbook["answers"]):
                old_q_id = row.get("question_id")
                new_q_id = q_id_map.get(int(old_q_id)) if old_q_id is not None else None
                if new_q_id is None:
                    stats["answers"]["skipped"] += 1
                    continue
                content = row.get("content")
                # Skip duplicate: same content + same question
                if content and Answer.objects.filter(
                    content=content, question_id=new_q_id
                ).exists():
                    stats["answers"]["skipped"] += 1
                    continue
                Answer.objects.create(
                    question_id=new_q_id,
                    content=content,
                    favorite=bool(row.get("favorite")),
                    status=row.get("status"),
                )
                stats["answers"]["created"] += 1

    wbook.close()
    return Response({"data": stats})
